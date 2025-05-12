import sys
from tkinter import messagebox
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox
from ui.ui_ana_ekran import Ui_MainWindow
from ui.ui_soru_ekleme import Ui_Form as Ui_SoruEkleme
from ui.ui_soru_yazdirma import Ui_Form as Ui_SoruYazdirma
from openpyxl import load_workbook
# filepath: [main.py](http://_vscodecontentref_/2)
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from PyQt5.QtGui import QTextDocument

import os
from openpyxl import Workbook, load_workbook

class SoruEklemeEkrani(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_SoruEkleme()
        self.ui.setupUi(self)
        self.ui.btnEkle.clicked.connect(self.soru_ekle)
        self.ui.btnKaydet.clicked.connect(self.kaydet)
        self.soruListesi = []

    def soru_ekle(self):
        soru = self.ui.textSoru.toPlainText()
        cevaplar = [
            self.ui.cevap1.text(),
            self.ui.cevap2.text(),
            self.ui.cevap3.text(),
            self.ui.cevap4.text(),
            self.ui.cevap5.text()
        ]
        dogru_index = -1
        for i, radio in enumerate([self.ui.radio1, self.ui.radio2, self.ui.radio3, self.ui.radio4, self.ui.radio5]):
            if radio.isChecked():
                dogru_index = i
                break

        if soru and all(cevaplar) and dogru_index != -1:
            self.soruListesi.append({
                "soru": soru,
                "cevaplar": cevaplar,
                "dogru": dogru_index
            })
            self.ui.textSoru.clear()
            for lineEdit in [self.ui.cevap1, self.ui.cevap2, self.ui.cevap3, self.ui.cevap4, self.ui.cevap5]:
                lineEdit.clear()
            for radio in [self.ui.radio1, self.ui.radio2, self.ui.radio3, self.ui.radio4, self.ui.radio5]:
                radio.setChecked(False)

    def kaydet(self):
        dosya_adi = "soru_bankasi.xlsx"  # Sabit dosya adı
        if os.path.exists(dosya_adi):
            wb = load_workbook(dosya_adi)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Soru", "Cevap A", "Cevap B", "Cevap C", "Cevap D", "Cevap E", "Doğru Şık"])
        if not self.soruListesi:
         QMessageBox.warning(self, "Uyarı", "Kaydedilecek soru yok!")
         return

        for s in self.soruListesi:
            dogru_harf = chr(65 + s['dogru'])  # 0->A, 1->B, ...
            ws.append([s['soru']] + s['cevaplar'] + [dogru_harf])

        wb.save(dosya_adi)
        self.soruListesi.clear()
        QMessageBox.information(self, "Başarılı", "Sorular başarıyla kaydedildi.")

class SoruSecmeEkrani(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_SoruYazdirma()
        self.ui.setupUi(self)
        self.ui.btnSec.clicked.connect(self.dosya_sec)
        self.ui.btnYazdir.clicked.connect(self.yazdir)
        self.icerik = ""

    def dosya_sec(self):
        dosya_adi, _ = QFileDialog.getOpenFileName(
            self, "Dosya Seç", "", "Excel Dosyaları (*.xlsx);;Text Dosyaları (*.txt)"
        )
        if dosya_adi:
            if dosya_adi.endswith('.xlsx'):
                wb = load_workbook(dosya_adi)
                ws = wb.active
                icerik = ""
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        continue  # başlık satırını atla
                    soru = row[0]
                    cevaplar = row[1:6]
                    dogru = row[6]
                    icerik += f"Soru: {soru}\n"
                    for idx, cevap in enumerate(cevaplar):
                        harf = chr(65 + idx)
                        dogru_mu = " (Doğru)" if harf == dogru else ""
                        icerik += f"  {harf}. {cevap}{dogru_mu}\n"
                    icerik += "\n"
                self.icerik = icerik
                self.ui.textAlan.setPlainText(self.icerik)
            else:
                with open(dosya_adi, "r", encoding="utf-8") as f:
                    self.icerik = f.read()
                    self.ui.textAlan.setPlainText(self.icerik)

    def yazdir(self):
        dosya_adi, _ = QFileDialog.getSaveFileName(self, "PDF Olarak Kaydet", "", "PDF Dosyası (*.pdf)")
        if not dosya_adi:
            return
        if not dosya_adi.endswith('.pdf'):
            dosya_adi += '.pdf'

        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(dosya_adi)

        doc = QTextDocument()
        doc.setPlainText(self.icerik)
        doc.print_(printer)

class AnaSayfa(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.btnYeniSoru.clicked.connect(self.yeni_soru_ekle)
        self.ui.btnSoruSec.clicked.connect(self.soru_sec)

    def yeni_soru_ekle(self):
        self.soruEkleme = SoruEklemeEkrani()
        self.soruEkleme.show()

    def soru_sec(self):
        self.soruSecme = SoruSecmeEkrani()
        self.soruSecme.show()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    pencere = AnaSayfa()
    pencere.show()
    sys.exit(app.exec_())